from bs4 import BeautifulSoup
from openpyxl import load_workbook
import plotly.graph_objs as go
import plotly.figure_factory as ff
import requests
import json
import webbrowser
import csv
import sqlite3
import time

CACHE_FILENAME = "covid_cache.json"
CACHE_DICT = {}
DB_NAME = "covid_usdaers.sqlite"

def build_county_url_dict():
    ''' Scrapes USDA ERS county-level datasets webpage and creates a dictionary for each dataset and its corresponding URL.

    PARAMETERS
    ----------
    none

    RETURNS
    -------
    dict:
        Dictionary of 4 county-level data sets available from the USDA ERS and their respective URLs.
    '''

    url = "https://www.ers.usda.gov/data-products/county-level-data-sets/"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    make_request_with_cache(url, soup.prettify())

    section = soup.find("div", style="margin-left: 4em;")
    indiv = section.find("ul")

    link_texts = []
    for words in indiv.find_all("li"):
        link_texts.append(words.text.strip())
    link_urls = []
    for items in indiv.find_all("a"):
        link_urls.append(items['data-id'])
    
    data_dict = {}
    for i in range(len(link_texts)):
        data_dict[link_texts[i]] = f"https://data.ers.usda.gov/reports.aspx?ID={link_urls[i]}"

    return data_dict

def npr_covid_data_dict():
    ''' Scrapes COVID-19 table on NPR webpage. Creates nested dictionary where each key has a dictionary value with "Cases" and "Deaths" as keys and numeric integers as values.
    
    PARAMETERS
    ----------
    none

    RETURNS
    -------
    dict:
        Nested dictionary with each key having a dictionary with 2 keys ("Cases", "Deaths").
        Example:
            {"United States": {"Cases": INT, "Deaths": INT}}
    '''

    npr_url = "https://apps.npr.org/dailygraphics/graphics/coronavirus-d3-us-map-20200312/table.html?initialWidth=1238&childId=responsive-embed-coronavirus-d3-us-map-20200312-table&parentTitle=Coronavirus%20Map%20And%20Graphics%3A%20Track%20The%20Spread%20In%20The%20U.S.%20%3A%20Shots%20-%20Health%20News%20%3A%20NPR&parentUrl=https%3A%2F%2Fwww.npr.org%2Fsections%2Fhealth-shots%2F2020%2F03%2F16%2F816707182%2Fmap-tracking-the-spread-of-the-coronavirus-in-the-u-s"
    npr_response = requests.get(npr_url)
    npr_soup = BeautifulSoup(npr_response.text, 'html.parser')

    covid_nums = {}

    names_list = []
    names = npr_soup.find_all("div", class_="cell cell-inner stateName")
    for n in names:
        names_list.append(n.text.strip())

    cases_list = []
    cases = npr_soup.find_all("div", class_="cell amt confirmed cell-inner")
    for c in cases:
        cases_list.append(clean_nums(c.text.strip()))

    deaths_list = []
    deaths = npr_soup.find_all("div", class_="cell amt deaths cell-inner")
    for d in deaths:
        deaths_list.append(clean_nums(d.text.strip()))

    for i in range(len(names_list)):
            covid_nums[names_list[i]] = {
                    'Cases': cases_list[i],
                    'Deaths': deaths_list[i]
                }
    
    return covid_nums

def npr_covid_data_time_pulled():
    ''' Scrapes COVID-19 table on NPR webpage to return the time and date of when the table was updated.
    
    PARAMETERS
    ----------
    none

    RETURNS
    -------
    str:
        A string showing time and date when NPR COVID-19 table was updated.
    '''

    npr_url = "https://apps.npr.org/dailygraphics/graphics/coronavirus-d3-us-map-20200312/table.html?initialWidth=1238&childId=responsive-embed-coronavirus-d3-us-map-20200312-table&parentTitle=Coronavirus%20Map%20And%20Graphics%3A%20Track%20The%20Spread%20In%20The%20U.S.%20%3A%20Shots%20-%20Health%20News%20%3A%20NPR&parentUrl=https%3A%2F%2Fwww.npr.org%2Fsections%2Fhealth-shots%2F2020%2F03%2F16%2F816707182%2Fmap-tracking-the-spread-of-the-coronavirus-in-the-u-s"
    npr_response = requests.get(npr_url)
    npr_soup = BeautifulSoup(npr_response.text, 'html.parser')

    time = []
    find_time = npr_soup.find("span", class_="latestTime")
    for ft in find_time:
        time.append(ft)
    
    return time[0]

def build_usda_ers_dict(dict1, dict2, dict3, dict4, dict5, dict6):
    ''' Combines 6 dictionaries into 1. Creates a nested dictionary where each key has a dictionary value with "Population", "Median Household Income", "Poverty Rate", "Unemployment Rate", "Completed HS Only Rate", and "College Completion Rate" as keys.
    
    PARAMETERS
    ----------
    dict1: dict
        A dictionary to be combined.
    
    dict2: dict
        A dictionary to be combined.
    
    dict3: dict
        A dictionary to be combined.
    
    dict4: dict
        A dictionary to be combined.
    
    dict5: dict
        A dictionary to be combined.
    
    dict6: dict
        A dictionary to be combined.

    RETURNS
    -------
    dict:
        Nested dictionary with each key having a dictionary with 6 keys ("Population", "Median Household Income", "Poverty Rate", "Unemployment Rate", "Completed HS Only Rate", "College Completion Rate").
        Example:
            {"Michigan": {"Population": INT, "Median Household Income": INT, "Poverty Rate": FLOAT, "Unemployment Rate": FLOAT, "Completed HS Only Rate": FLOAT, "College Completion Rate": FLOAT}}
    '''

    socioecon = {**dict1, **dict2, **dict3, **dict4, **dict5, **dict6}

    for key, value in socioecon.items():
        if key in dict1 and key in dict2 and key in dict3 and key in dict4 and key in dict5 and key in dict6:
            socioecon[key] = [
                value,
                dict1[key],
                dict2[key],
                dict3[key],
                dict4[key],
                dict5[key]
            ]
    
    socioecon2 = {}

    for k, items in socioecon.items():
        socioecon2[k] = {
            'Population': items[1]['Population'],
            'Median Household Income': items[0]['Median Household Income'],
            'Poverty Rate': items[2]['Poverty Rate'],
            'Unemployment Rate': items[5]['Unemployment Rate'],
            "Completed HS Only Rate": items[3]["Completed HS Only Rate"],
            "College Completion Rate": items[4]["College Completion Rate"],

        }

    return socioecon2

def build_socioecon_dict(names, data, key):
    ''' Takes in a list of names and values to create a nested dictionary where each key has a dictionary value with 'key' parameter as the key and 'data' parameter as the value.
    
    PARAMETERS
    ----------
    names: list
        A list of names to be used as keys.
    
    data: list
        A list of values (INTs, FLOATs) to be used as values.

    key: str
        A string to be used as a key in the dictionary value.

    RETURNS
    -------
    dict:
        Nested dictionary with each key having a dictionary with 1 key ('key' parameter).
        Example:
        {"name": {"key": data}}
    '''

    socioecon = {}
    for i in range(len(names)):
        socioecon[names[i]] = {
            key: data[i]
        }
    
    return socioecon

def get_excel_data(workbook, sheet, cellrange):
    ''' Opens an Excel workbook. Reads data in specified worksheet and cell range. Returns the data as a list.
    
    PARAMETERS
    ----------
    workbook: str
        The name of a workbook to be accessed.
    
    sheet: str
        The name of a sheet within the workbook.

    cellrange: str
        The cell range of the data to be accessed.

    RETURNS
    -------
    list:
        The data returned from the cell range specified.
    '''

    data = []

    wb = load_workbook(workbook)
    ws = wb[sheet]
    rows = ws[cellrange]

    for r in rows:
        for cells in r:
            data.append(cells.value)

    return data

def create_database():
    ''' Creates a SQL database with 3 tables: "CovidCounty", "CovidState", "SocioeconomicStates". 
    
    PARAMETERS
    ----------
    none

    RETURNS
    -------
    none
    '''

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    drop_county_covid_sql = "DROP TABLE IF EXISTS 'CovidCounty'"
    drop_state_covid_sql = "DROP TABLE IF EXISTS 'CovidState'"
    drop_states_usda_sql = "DROP TABLE IF EXISTS 'SocioeconomicStates'"
    drop_mi_usda_sql = "DROP TABLE IF EXISTS 'SocioeconomicMichigan'"

    create_county_covid_sql = '''
        CREATE TABLE IF NOT EXISTS "CovidCounty" (
            "Id" INTEGER PRIMARY KEY AUTOINCREMENT,
            "Date" TEXT NOT NULL,
            "County" TEXT NOT NULL,
            "StateName" TEXT NOT NULL,
            "Fips" INTEGER NOT NULL,
            "CountyCases" INTEGER,
            "CountyDeaths" INTEGER
        )
    '''

    create_state_covid_sql = '''
        CREATE TABLE IF NOT EXISTS "CovidState" (
            "Id" INTEGER PRIMARY KEY AUTOINCREMENT,
            "Name" TEXT NOT NULL,
            "StateCases" INTEGER NOT NULL,
            "StateDeaths" INTEGER NOT NULL
        )
    '''

    create_states_usda_sql = '''
        CREATE TABLE IF NOT EXISTS "SocioeconomicStates" (
            "Id" INTEGER PRIMARY KEY AUTOINCREMENT,
            "StateName" TEXT NOT NULL,
            "StatePopulation" INTEGER NOT NULL,
            "StateMedianIncome" INTEGER NOT NULL,
            "StatePovertyRate" DECIMAL NOT NULL,
            "StateUnemploymentRate" DECIMAL NOT NULL,
            "StateCompHSOnlyRate" DECIMAL NOT NULL,
            "StateCompCollRate" DECIMAL NOT NULL
        )
    '''

    cur.execute(drop_county_covid_sql)
    cur.execute(drop_state_covid_sql)
    cur.execute(drop_states_usda_sql)
    cur.execute(drop_mi_usda_sql)
    cur.execute(create_county_covid_sql)
    cur.execute(create_state_covid_sql)
    cur.execute(create_states_usda_sql)

    conn.commit()
    conn.close()

def populate_database():
    ''' Populates 3 tables in SQL database with data from a variety of sources.
    
    PARAMETERS
    ----------

    RETURNS
    -------
    '''

    data_header = []
    data_rows = []

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    with open("covid_data/us-counties.csv", 'r') as csvfile:
        data = []
        csv_header = csv.reader(csvfile)
        for h in csv_header:
            data.append(h)
        data_header.extend(data[0])
        data_rows.extend(data[1:])

    insert_county_covid_sql = '''
        INSERT INTO CovidCounty
        VALUES (NULL, ?, ? , ?, ?, ?, ?)
    '''

    for dr in data_rows:
        cur.execute(insert_county_covid_sql, [
            dr[0],
            dr[1],
            dr[2],
            dr[3],
            dr[4],
            dr[5]
        ])

    insert_state_covid_sql = '''
        INSERT INTO CovidState
        VALUES (NULL, ?, ?, ?)
    '''

    for k,v in npr_covid_data_dict().items():
        cur.execute(insert_state_covid_sql, [
            k,
            v['Cases'],
            v['Deaths']
        ])
    
    insert_state_ers_sql = '''
        INSERT INTO SocioeconomicStates
        VALUES (Null, ?, ?, ?, ?, ?, ?, ?)
    '''
    
    with open("USDA_ERS_Data.json") as file_obj:
        for k, v in json.load(file_obj).items():
            cur.execute(insert_state_ers_sql, [
                k,
                v['Population'],
                v['Median Household Income'],
                v['Poverty Rate'],
                v['Unemployment Rate'],
                v['Completed HS Only Rate'],
                v['College Completion Rate']
            ])

    conn.commit()
    conn.close()

def clean_county_covid_data():
    ''' Reads in COVID-19 CSV data, cleans it by converting numeric string data into numeric data, and then creates a nested dictionary.
    
    PARAMETERS
    ----------
    none

    RETURNS
    -------
    dict:
        A nested dictionary for each state with county names and county COVID-19 data as values.
        Example:
            {"Michigan": {"Washtenaw": {"Cases": INT, "Deaths": INT}}}
    '''

    data_header = []
    data_rows = []
    county_dict = {}

    with open("covid_data/us-counties.csv", 'r') as csvfile:
        data = []
        csv_header = csv.reader(csvfile)
        for h in csv_header:
            data.append(h)
        data_header.extend(data[0])
        data_rows.extend(data[1:])

    for dr in data_rows:
        if dr[2] not in county_dict:
            county_dict[dr[2]] = {
                dr[1]: {
                    "Cases": int(dr[4]),
                    "Deaths": int(dr[5])
                }
            }
        elif dr[2] in county_dict:
            county_dict[dr[2]].update(
            {
                dr[1]: {
                    "Cases": int(dr[4]),
                    "Deaths": int(dr[5])
                }
            }
            )

    return county_dict

def clean_nums(data):
    ''' Takes in a value and removes a comma in order to convert the value into an integer.
    
    PARAMETERS
    ----------
    data: str
        A value to be cleaned.

    RETURNS
    -------
    int:
        The value in integer form.
    '''

    data = data.replace(',','')
    return int(data)

def income_to_int(values):
    ''' Takes in a string of income data and removes both "$" and "," to produce a clean integer.
    
    PARAMETERS
    ----------
    values: list
        A list of income values to be cleaned.

    RETURNS
    -------
    list:
        A cleaned list of the income values in integer form.
    '''

    list_of_values = []

    for v in values: 
        v = v.replace("$",'')
        v = v.replace(',','')
        list_of_values.append(int(v))

    return list_of_values

def convert_to_percent(values):
    ''' Attempts to convert a list of values into a float in percentage format.
    
    PARAMETERS
    ----------
    values: list
        A list of values to be converted into float percentage format.

    RETURNS
    list:
        The list of values where the value are either in float percentage format or their original form.
    -------
    '''

    list_of_values = []
    
    for v in values:
        try:
            percent = float('{0:.2f}'.format((v * 100)))
            list_of_values.append(percent)
        except:
            list_of_values.append(v)
    
    return list_of_values

def write_to_json(filename, data):
    ''' Takes in data and writes it out into JSON format.
    
    PARAMETERS
    ----------
    filename: str
        The name of the JSON file to be created.
    
    data: dict
        The data to be written into the JSON file.

    RETURNS
    -------
    none
    '''

    with open(filename, "w") as file_obj:
        json.dump(data, file_obj, indent=4)

def clean_excel_data():
    ''' Calls on various functions to access and clean XLSX data. Build dictionaries using XLSX data and then writes that data to JSON file.
    
    PARAMETERS
    ----------
    none

    RETURNS
    -------
    none
    '''

    # getting state socioeconomic data
    comp_coll_names = get_excel_data("socioeconomic_data/EducationReportCompColl.xlsx", "EducationReport", 'A6:A56')
    comp_coll_perc = get_excel_data("socioeconomic_data/EducationReportCompColl.xlsx", "EducationReport", 'F6:F56')

    comp_hs_only_names = get_excel_data("socioeconomic_data/EducationReportHSOnly.xlsx", "EducationReport", 'A6:A56')
    comp_hs_only = get_excel_data("socioeconomic_data/EducationReportHSOnly.xlsx", "EducationReport", 'F6:F56')

    pop_names = get_excel_data("socioeconomic_data/PopulationReport.xlsx", "PopulationReport", 'A6:A56')
    pop_num = get_excel_data("socioeconomic_data/PopulationReport.xlsx", "PopulationReport", 'E6:E56')

    poverty_names = get_excel_data("socioeconomic_data/PovertyReportPercent.xlsx", "PovertyReport", 'A7:A57')
    poverty_perc = get_excel_data("socioeconomic_data/PovertyReportPercent.xlsx", "PovertyReport", 'E7:E57')

    unemp_names = get_excel_data("socioeconomic_data/UnemploymentReportPercent.xlsx", "UnemploymentReport", 'B4:B54')
    unemp_perc = get_excel_data("socioeconomic_data/UnemploymentReportPercent.xlsx", "UnemploymentReport", 'K4:K54')

    med_income_names = get_excel_data("socioeconomic_data/UnemploymentReportPercent.xlsx", "UnemploymentReport", 'B4:B54')
    med_income = get_excel_data("socioeconomic_data/UnemploymentReportPercent.xlsx", "UnemploymentReport", 'L4:L54')

    # building state socioeconomic dictionaries
    comp_coll_dict = build_socioecon_dict(comp_coll_names, convert_to_percent(comp_coll_perc), "College Completion Rate")
    comp_hs_only_dict = build_socioecon_dict(comp_hs_only_names, convert_to_percent(comp_hs_only), "Completed HS Only Rate")
    poverty_dict = build_socioecon_dict(poverty_names, poverty_perc, "Poverty Rate")
    pop_dict = build_socioecon_dict(pop_names, pop_num, "Population")
    unemp_dict = build_socioecon_dict(unemp_names, unemp_perc, "Unemployment Rate")
    med_income_dict = build_socioecon_dict(med_income_names, income_to_int(med_income), "Median Household Income")
    
    usda_ers_data = build_usda_ers_dict(pop_dict, poverty_dict, comp_hs_only_dict, comp_coll_dict, unemp_dict, med_income_dict)

    # writing data to json
    write_to_json("USDA_ERS_Data.json", usda_ers_data)

def access_state_sql_database(state):
    ''' Makes a request to SQL database to access state-specific information on COVID-19 data and returns it as a list.
    
    PARAMETERS
    ----------
    state: str
        The state for which the user would like to see data on.

    RETURNS
    -------
    list:
        The results of the SQL query.
    '''

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    query = f'''
        SELECT StateName, County, MAX(CountyCases), MAX(CountyDeaths)
        FROM CovidCounty
        WHERE StateName = "{state}"
        GROUP BY County
        ORDER BY MAX(CountyCases) DESC
    '''
    result = cur.execute(query).fetchall()
    conn.close()
    return result

def access_national_sql_database():
    ''' Makes a request to SQL database to access state information on COVID-19 data, USDA ERS socioeconomic data for each state, and returns it as a list.
    
    PARAMETERS
    ----------
    none

    RETURNS
    -------
    list:
        The results of the SQL query.
    '''

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    query = '''
        SELECT Name, MAX(StateCases), MAX(StateDeaths), ss.StatePopulation, ss.StateMedianIncome, ss.StateUnemploymentRate, ss.StatePovertyRate, ss.StateCompCollRate, ss.StateCompHSOnlyRate
        FROM CovidState
            JOIN SocioeconomicStates as ss
            ON CovidState.Name = ss.StateName
        GROUP BY Name
        ORDER BY MAX(StateCases) DESC
    '''
    result = cur.execute(query).fetchall()
    conn.close()
    return result

def create_and_show_figures(user_input):
    ''' Using Plotly, creates a bar graph and a table based on user_input value. Launches the visuals in the user's browser.
    
    PARAMETERS
    ----------
    user_input: str
        The information the user would like to see presented in visual form.

    RETURNS
    -------
    none
    '''

    if user_input == "nation":
        state_names = []
        state_cases = []
        state_deaths = []
        table_data = [["State", "Cases", "Deaths", "Population", "Median Income", "Unemployment Rate", "Poverty Rate", "College Completion Rate", "Completed High School Only Rate"]]
        for data in access_national_sql_database():
            table_data.append([data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8]])
            state_names.append(data[0])
            state_cases.append(data[1])
            state_deaths.append(data[2])
        
        trace1 = go.Bar(name="Cases", x=state_names, y=state_cases, xaxis="x2", yaxis="y2")
        trace2 = go.Bar(name="Deaths", x=state_names, y=state_deaths, xaxis="x2", yaxis="y2")
    else:
        county_names = []
        county_cases = []
        county_deaths = []
        state_socio = []
        table_data = [['County', 'Cases', 'Deaths']]
        for data in access_state_sql_database(user_input):
            table_data.append([data[1], data[2], data[3]])
            county_names.append(data[1])
            county_cases.append(data[2])
            county_deaths.append(data[3])

        for national_data in access_national_sql_database():
            if national_data[0] == user_input:
                state_socio.extend([national_data[0], national_data[3], national_data[4], national_data[5], national_data[6], national_data[7], national_data[8]])

        print(f"\nHere is socioeconomic data for {user_input}:")
        time.sleep(1)
        print(f"\nPopulation: {state_socio[1]}")
        time.sleep(1)
        print(f"Median Household Income: {state_socio[2]}")
        time.sleep(1)
        print(f"Unemployment Rate: {state_socio[3]}")
        time.sleep(1)
        print(f"Poverty Rate: {state_socio[4]}")
        time.sleep(1)
        print(f"College Completion Rate: {state_socio[5]}")
        time.sleep(1)
        print(f"Completed High School Only Rate: {state_socio[6]}")

        trace1 = go.Bar(name="Cases", x=county_names, y=county_cases, xaxis="x2", yaxis="y2")
        trace2 = go.Bar(name="Deaths", x=county_names, y=county_deaths, xaxis="x2", yaxis="y2")
    
    table = ff.create_table(table_data)

    table.add_traces([trace1, trace2])

    table["layout"]["xaxis2"] = {}
    table["layout"]["yaxis2"] = {}

    table.layout.yaxis.update({"domain":[0, .45]})
    table.layout.yaxis2.update({"domain":[.6, 1]})

    table.layout.yaxis2.update({"anchor":"x2"})
    table.layout.xaxis2.update({"anchor":"y2"})
    table.layout.yaxis2.update({"title":"COVID-19"})

    table.layout.margin.update({"t":75, "l":50})

    if user_input == "nation":
        table.layout.update({"title":"National 2020 COVID-19 Numbers"})
    else:
        table.layout.update({"title":f"{user_input} 2020 COVID-19 Numbers"})

    print("\nThe visuals will now launch in your browswer.")
    time.sleep(2)
    table.show()

def open_cache():
    ''' Opens the cache file if it exists and loads the JSON into
    the CACHE_DICT dictionary.
    if the cache file doesn't exist, creates a new cache dictionary
    
    Parameters
    ----------
    None
    
    Returns
    -------
    dict:
        The opened cache
    '''
    try:
        cache_file = open(CACHE_FILENAME, 'r')
        cache_contents = cache_file.read()
        cache_dict = json.loads(cache_contents)
        cache_file.close()
    except:
        cache_dict = {}
    return cache_dict

def save_cache(cache_dict):
    ''' Saves the current state of the cache to disk
    
    Parameters
    ----------
    cache_dict: dict
        The dictionary to save
    
    Returns
    -------
    None
    '''
    dumped_json_cache = json.dumps(cache_dict)
    fw = open(CACHE_FILENAME,"w")
    fw.write(dumped_json_cache)
    fw.close() 

def make_request_with_cache(cache_key, cache_value):
    '''Check the cache for a saved result for this cache_key:cache_value
    combo. If the result is found, return it. Otherwise send a new 
    request, save it, then return it.
    
    Parameters
    ----------
    cache_key: string
        Various strings to be used as keys in CACHE_DICT
    cache_value: string
        Information to be saved as the value in CACHE_DICT
    
    Returns
    -------
    dict
        the results of the query as a dictionary loaded from cache
        JSON
    '''
    if cache_key in CACHE_DICT.keys():
        # print("Using cache")
        return CACHE_DICT[cache_key]
    else:
        # print("Fetching")
        CACHE_DICT[cache_key] = cache_value
        save_cache(CACHE_DICT)
        return CACHE_DICT[cache_key]

if __name__ == "__main__":
    CACHE_DICT = open_cache()
    STATES = ["Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado", "Connecticut", "Delaware", "District of Columbia", "Florida", "Georgia", "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota", "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada", "New Hampshire", "New Jersey", "New Mexico", "New York", "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon", "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington", "West Virginia", "Wisconsin", "Wyoming"]
    STATE_INPUT_NUM = None
    URL_LIST = []

    clean_excel_data()
    create_database()
    populate_database()
    write_to_json("US_Covid.json", npr_covid_data_dict())
    write_to_json("County_Covid.json", clean_county_covid_data())

    welcome_message = '''
    Welcome!\n
    This interactive program combines USDA Economic Research Service (USDA ERS) and COVID-19 data.\n
    You will be able to see both state- and county-level COVID-19 data.\n
    State-level COVID-19 data comes from scraping NPR's COVID-19 tracking webpage. The webpage is updated periodically and a timestamp is provided.\n
    County-level COVID-19 data comes from the New York Times' GitHub repository. This data was downloaded as a CSV file on April 27th, with numbers current as of April 26th.\n
    If you choose to see COVID-19 data for a specific state, then you will also be able to see a number of socioeconomic data harvested from USDA ERS data about that specific state.\n
    The socioeconomic data are:\n
    - Population\n
    - Median Household Income\n
    - Unemployment Rate\n
    - Poverty Rate\n
    - College Completion Rate\n
    - High School Only Completion Rate\n
    Both national and state COVID-19 data will be presented in bar and table form if you like. If you select a state, the socioeconomic data will print to your terminal and the COVID-19 data will launch in your browser.\n
    '''

    for wm in welcome_message.split("\n"):
        if wm == '':
            print(wm)
        if wm != '':
            print(wm)
            time.sleep(3)
    
    print("First, let's begin with the USDA ERS data. Here are the data sets being used:\n")
    time.sleep(2.5)

    change = True
    while True:
        while change is True:
            counter = 1
            for k,v in build_county_url_dict().items():
                print(f"[{counter}] {k}")
                URL_LIST.append(v)
                counter += 1
                time.sleep(.5)

            webpage = input(f"\nChoose a number to launch the webpage for the respective dataset, 'next' to see COVID-19 data, or 'exit' to leave this program:\n")

            if webpage.lower() == "exit":
                exit()

            elif webpage.isnumeric():
                webpage_num = int(webpage)
                if 1 <= webpage_num <= 4:
                    for i in range(len(URL_LIST)):
                        webbrowser.open((URL_LIST[webpage_num - 1]))
                else:
                    print("Please choose a number between 1 and 4.\n")

            elif webpage.lower() == "next":
                switch = True
                while switch is True:
                    time.sleep(1)
                    covid_data = input("\nYou can see COVID-19 data for the entire nation or a specific state. Enter 'nation', 'state', 'back' to go back and view USDA ERS data, or 'exit'.\n")

                    if covid_data.lower() == "exit":
                        exit()
                    
                    elif covid_data.lower() == "back":
                        switch = False
                        change = True

                    elif covid_data.lower() == "nation":
                        print(f"\nThis data is accurate as of {npr_covid_data_time_pulled()}.\n")
                        for k,v in npr_covid_data_dict().items():
                            print(f"{k}: Cases - {v['Cases']} | Deaths - {v['Deaths']}")
                            time.sleep(.3)

                        visuals = input("\nThis data can be presented visually. The COVID-19 data will be presented in both bar graph and table form. The socioeconimc data will be presented in table form only. Would you like to see it? Enter 'yes', 'back', or 'exit'.\n")
                        
                        if visuals.lower() == "exit":
                            exit()
                        
                        elif visuals.lower() == "yes":
                            i = 0
                            create_and_show_figures(covid_data.lower())
                            while i == 0:
                                user_input = input("\nEnter 'back' for the previous prompt or 'exit' to leave the program.\n")
                                
                                if user_input.lower() == "exit":
                                    exit()
                                
                                elif user_input.lower() == "back":
                                    i = 1
                                    switch = True
                                
                                else:
                                    print("Not a valid entry. Try again.\n")
                        
                        elif visuals.lower() == "back":
                            switch = True
                        
                        else:
                            print("Not a valid entry. Try again.\n")
                    
                    elif covid_data.lower() == "state":
                        turn = True
                        while turn is True:
                            counter = 1
                            for s in STATES:
                                print(f"[{counter}] {s}")
                                counter += 1
                                time.sleep(.2)
                            
                            state_data = input(f"\nChoose a number to see COVID-19 for a specific state (by county), 'back', or 'exit' to leave this program.\n")

                            if state_data.lower() == "exit":
                                exit()
                            
                            elif state_data.lower() == "back":
                                turn = False
                                switch = True
                            
                            elif state_data.isnumeric():
                                STATE_INPUT_NUM = int(state_data)
                                
                                if 1 <= STATE_INPUT_NUM <= 51:
                                    time.sleep(1)
                                    for i in range(len(STATES)):
                                        print(f"\nHere is the data for {STATES[STATE_INPUT_NUM - 1]}. It is accurate as of April 26th.\n")
                                        time.sleep(1)
                                        for data in access_state_sql_database(STATES[STATE_INPUT_NUM - 1]):
                                            print(f"{data[1]}: Cases - {data[2]} | Deaths - {data[3]}")
                                            time.sleep(.3)
                                        break
                                
                                else:
                                    print("Please choose a number between 1 and 51.\n")
                            
                                visuals = input("\nThis data can be presented visually. The COVID-19 data will be presented in both bar graph and table form. The socioeconimc data will be presented in table form only and will print out to your terminal. Would you like to see it? Enter 'yes', 'back', or 'exit'.\n")

                                if visuals.lower() == "exit":
                                    exit()

                                elif visuals.lower() == "yes":
                                    time.sleep(1)
                                    create_and_show_figures(STATES[STATE_INPUT_NUM - 1])
                                    state_input = input("\nEnter 'back' to see COVID-19 data for another state or 'exit' to leave the program.\n")
                                    
                                    if state_input.lower() == "exit":
                                        exit()
                                    
                                    elif state_input.lower() == "back":
                                        turn = True

                                elif visuals.lower() == "back":
                                    turn = True
                                    switch = True
                                
                                else:
                                    print("Not a valid entry. Try again.\n")

                            else:
                                print("Not a valid entry. Try again.\n")

                    else:
                        print("Not a valid entry. Try again.\n")

            else:
                print("Not a valid entry. Try again.\n")